# -*- coding: utf-8 -*-
import io
import os
import re
import random
from datetime import datetime
from typing import Dict, Tuple, List, Set

import requests
import streamlit as st
import openpyxl
from dateutil.relativedelta import relativedelta
from docx import Document
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

st.set_page_config(
    page_title="Εργαλείο Κατανομής Ανθρωπομηνών",
    page_icon="📊",
    layout="wide",
)

# ------------------------- GitHub assets -------------------------
TEMPLATE_URL = "https://raw.githubusercontent.com/dimitrisaronis1-dev/MANMONTHS-8/main/AM%20TEST%201.xlsx"
LOGO_URL = "https://raw.githubusercontent.com/dimitrisaronis1-dev/MANMONTHS-8/main/SPACE%20LOGO_colored%20horizontal.png"
GUIDE_DOC_URL = "https://raw.githubusercontent.com/dimitrisaronis1-dev/MANMONTHS-8/main/%CE%9F%CE%B4%CE%B7%CE%B3%CE%AF%CE%B5%CF%82%20%CF%87%CF%81%CE%AE%CF%83%CE%B7%CF%82%20%CE%95%CF%81%CE%B3%CE%B1%CE%BB%CE%B5%CE%AF%CE%BF%CF%85%20%CE%9A%CE%B1%CF%84%CE%B1%CE%BD%CE%BF%CE%BC%CE%AE%CF%82%20%CE%91%CE%9C.docx"
LOGO_WIDTH_PX = 380  # ~10cm

# ------------------------- Styles -------------------------
st.markdown(
    """
<style>
.block-container { padding-top: 1.6rem; padding-bottom: 2rem; }
h1, h2, h3 { letter-spacing: -0.02em; }

.notice {
  border-left: 4px solid #4C78A8;
  background: rgba(76,120,168,0.08);
  padding: 10px 12px;
  border-radius: 10px;
}
.successbox {
  border-left: 4px solid #2E7D32;
  background: rgba(46,125,50,0.08);
  padding: 10px 12px;
  border-radius: 10px;
}
.errorbox {
  border-left: 4px solid #C62828;
  background: rgba(198,40,40,0.08);
  padding: 10px 12px;
  border-radius: 10px;
}
hr { margin: 1.2rem 0 1.2rem 0; }

/* Sticky bottom controls in sidebar */
section[data-testid="stSidebar"] > div {
  display: flex;
  flex-direction: column;
  height: 100%;
}
.sidebar-bottom {
  margin-top: auto;
  padding-top: 12px;
  padding-bottom: 6px;
}
</style>
""",
    unsafe_allow_html=True,
)

# ------------------------- Excel settings & styles -------------------------
MAX_YEARLY_CAPACITY = 11
YELLOW_RGB_CANDIDATES = {"FFFF00", "FFFFFF00"}

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _http_get_bytes(url: str, timeout: int = 30) -> bytes:
    r = requests.get(url, timeout=timeout)
    r.raise_for_status()
    return r.content


@st.cache_data(show_spinner=False)
def get_template_bytes() -> bytes:
    return _http_get_bytes(TEMPLATE_URL)


@st.cache_data(show_spinner=False)
def get_logo_bytes() -> bytes:
    return _http_get_bytes(LOGO_URL)


@st.cache_data(show_spinner=False)
def get_guide_doc_bytes() -> bytes:
    return _http_get_bytes(GUIDE_DOC_URL)


@st.cache_data(show_spinner=False)
def get_guide_text() -> str:
    """Extract plain text from the DOCX guide."""
    doc_bytes = get_guide_doc_bytes()
    doc = Document(io.BytesIO(doc_bytes))
    paras = []
    for p in doc.paragraphs:
        t = (p.text or "").strip()
        if t:
            paras.append(t)
    return "\n\n".join(paras).strip()


def norm_period(s: str) -> str:
    return str(s).replace("—", "-").replace("–", "-").replace("  ", " ").strip()


def parse_date(text: str, is_start: bool = True) -> datetime:
    t = str(text).strip()
    if "σήμερα" in t.lower() or "simera" in t.lower():
        if not is_start:
            return datetime.today()
        raise ValueError("Το 'Σήμερα' επιτρέπεται μόνο ως ημερομηνία λήξης.")

    if re.match(r"^\d{4}$", t):
        return datetime.strptime(("01/01/" if is_start else "31/12/") + t, "%d/%m/%Y")

    if re.match(r"^\d{1,2}/\d{4}$", t):
        if is_start:
            return datetime.strptime("01/" + t, "%d/%m/%Y")
        d = datetime.strptime("01/" + t, "%d/%m/%Y")
        return d + relativedelta(months=1) - relativedelta(days=1)

    if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", t):
        return datetime.strptime(t, "%d/%m/%Y")

    raise ValueError(f"Unsupported date format: {t}")


def parse_period(p: str) -> Tuple[datetime, datetime]:
    p = norm_period(p)
    if re.match(r"^\d{4}$", p):
        return parse_date(p, True), parse_date(p, False)
    parts = [x.strip() for x in p.split("-")]
    if len(parts) != 2:
        raise ValueError(f"Invalid period: {p}")
    return parse_date(parts[0], True), parse_date(parts[1], False)


def month_range(start: datetime, end: datetime) -> List[Tuple[int, int]]:
    cur = datetime(start.year, start.month, 1)
    endm = datetime(end.year, end.month, 1)
    out: List[Tuple[int, int]] = []
    while cur <= endm:
        out.append((cur.year, cur.month))
        cur += relativedelta(months=1)
    return out


def is_light_color(hex_color: str) -> bool:
    hex_color = hex_color.lstrip("#")
    rgb = tuple(int(hex_color[i : i + 2], 16) for i in (0, 2, 4))
    luminance = (0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]) / 255
    return luminance > 0.5


class Model:
    def __init__(self):
        self.ym_to_col: Dict[Tuple[int, int], int] = {}
        self.col_to_ym: Dict[int, Tuple[int, int]] = {}
        self.year_blocks: List[Tuple[int, int, int]] = []

        self.row_period: Dict[int, str] = {}
        self.row_months: Dict[int, Set[Tuple[int, int]]] = {}
        self.row_months_list: Dict[int, List[Tuple[int, int]]] = {}
        self.row_requested_am: Dict[int, int] = {}
        self.row_is_yellow: Dict[int, bool] = {}
        self.row_project_index: Dict[int, int] = {}

        self.month_owner: Dict[Tuple[int, int], int] = {}
        self.yearly_totals: Dict[int, int] = {}
        self.row_x_count: Dict[int, int] = {}

    def is_yellow_row(self, r: int) -> bool:
        return bool(self.row_is_yellow.get(r, False))

    def donor_can_deallocate(self, r: int) -> bool:
        return self.row_x_count.get(r, 0) > 1


def build_headers_and_maps(ws, years: List[int], start_col: int, year_row: int, month_row: int) -> Tuple[Dict[Tuple[int, int], int], int]:
    col = start_col
    month_col_map: Dict[Tuple[int, int], int] = {}

    for y in years:
        year_start = col
        year_cell = ws.cell(year_row, col)

        rc = lambda: random.randint(0, 255)
        rand_hex = "%02X%02X%02X" % (rc(), rc(), rc())
        year_cell.fill = PatternFill(start_color=rand_hex, end_color=rand_hex, fill_type="solid")
        year_cell.font = Font(color="FFFFFF" if not is_light_color(rand_hex) else "000000")

        for m in range(1, 13):
            ws.cell(month_row, col).value = m
            month_col_map[(y, m)] = col
            col += 1

        ws.merge_cells(start_row=year_row, start_column=year_start, end_row=year_row, end_column=col - 1)
        year_cell.value = y

    for c in range(start_col, col):
        ws.cell(year_row, c).border = thin_border
        ws.cell(month_row, c).border = thin_border

    return month_col_map, col


def move_x(model: Model, ws, r: int, from_ym: Tuple[int, int], to_ym: Tuple[int, int]) -> bool:
    if from_ym not in model.ym_to_col or to_ym not in model.ym_to_col:
        return False
    if ws.cell(r, model.ym_to_col[from_ym]).value != "X":
        return False
    if to_ym not in model.row_months.get(r, set()):
        return False
    if to_ym in model.month_owner:
        return False
    if to_ym[0] != from_ym[0] and model.yearly_totals.get(to_ym[0], 0) >= MAX_YEARLY_CAPACITY:
        return False

    ws.cell(r, model.ym_to_col[from_ym]).value = None
    model.month_owner.pop(from_ym, None)
    model.yearly_totals[from_ym[0]] = model.yearly_totals.get(from_ym[0], 0) - 1
    model.row_x_count[r] = model.row_x_count.get(r, 0) - 1

    ws.cell(r, model.ym_to_col[to_ym]).value = "X"
    model.month_owner[to_ym] = r
    model.yearly_totals[to_ym[0]] = model.yearly_totals.get(to_ym[0], 0) + 1
    model.row_x_count[r] = model.row_x_count.get(r, 0) + 1
    return True


def add_x(model: Model, ws, r: int, ym: Tuple[int, int]) -> bool:
    if ym not in model.row_months.get(r, set()):
        return False
    if ym in model.month_owner:
        return False
    if model.yearly_totals.get(ym[0], 0) >= MAX_YEARLY_CAPACITY:
        return False

    ws.cell(r, model.ym_to_col[ym]).value = "X"
    model.month_owner[ym] = r
    model.yearly_totals[ym[0]] = model.yearly_totals.get(ym[0], 0) + 1
    model.row_x_count[r] = model.row_x_count.get(r, 0) + 1
    return True


def free_capacity_in_year(model: Model, ws, year: int) -> bool:
    allocs = [(ym, r) for ym, r in model.month_owner.items() if ym[0] == year]
    allocs.sort(key=lambda t: (model.is_yellow_row(t[1]), model.row_x_count.get(t[1], 0)))

    for allow_yellow in (False, True):
        for ym_from, r in allocs:
            if (not allow_yellow) and model.is_yellow_row(r):
                continue
            for ym_to in model.row_months_list.get(r, []):
                if ym_to[0] == year:
                    continue
                if ym_to in model.month_owner:
                    continue
                if model.yearly_totals.get(ym_to[0], 0) >= MAX_YEARLY_CAPACITY:
                    continue
                if move_x(model, ws, r, ym_from, ym_to):
                    return True

    allocs = [(ym, r) for ym, r in model.month_owner.items() if ym[0] == year]
    allocs.sort(key=lambda t: (model.is_yellow_row(t[1]), model.row_x_count.get(t[1], 0)))
    for ym_from, r in allocs:
        if model.is_yellow_row(r):
            continue
        if not model.donor_can_deallocate(r):
            continue
        ws.cell(r, model.ym_to_col[ym_from]).value = None
        model.month_owner.pop(ym_from, None)
        model.yearly_totals[year] = model.yearly_totals.get(year, 0) - 1
        model.row_x_count[r] = model.row_x_count.get(r, 0) - 1
        return True

    return False


def make_month_free(model: Model, ws, ym: Tuple[int, int], depth: int = 0, max_depth: int = 6, visited=None) -> bool:
    if visited is None:
        visited = set()
    if ym in visited:
        return False
    visited.add(ym)

    occ_r = model.month_owner.get(ym)
    if occ_r is None:
        return True

    from_ym = ym
    dests_same = [d for d in model.row_months_list.get(occ_r, []) if d != from_ym and d[0] == from_ym[0]]
    dests_other = [d for d in model.row_months_list.get(occ_r, []) if d != from_ym and d[0] != from_ym[0]]

    for to_ym in dests_same + dests_other:
        if to_ym in model.month_owner:
            if depth >= max_depth:
                continue
            if not make_month_free(model, ws, to_ym, depth + 1, max_depth, visited):
                continue

        if to_ym[0] != from_ym[0] and model.yearly_totals.get(to_ym[0], 0) >= MAX_YEARLY_CAPACITY:
            if not free_capacity_in_year(model, ws, to_ym[0]):
                continue

        if move_x(model, ws, occ_r, from_ym, to_ym):
            return True

    return False


def compute_unallocated_reasons(model: Model, target_row: int) -> str:
    requested = model.row_requested_am.get(target_row, 0)
    allocated = model.row_x_count.get(target_row, 0)
    if requested <= 0 or allocated >= requested:
        return ""

    reasons = set()
    months = model.row_months_list.get(target_row, [])
    for (y, m) in months:
        if (y, m) in model.month_owner and model.month_owner[(y, m)] != target_row:
            other_row = model.month_owner[(y, m)]
            other_idx = model.row_project_index.get(other_row, "?")
            reasons.add(f"Month {m}/{y} already allocated by Project {other_idx}")
        else:
            if model.yearly_totals.get(y, 0) >= MAX_YEARLY_CAPACITY:
                reasons.add(f"Year {y} capacity reached")

    if not reasons:
        return "Capacity/Month taken by other projects."
    ordered = sorted(reasons, key=lambda s: (s.startswith("Year "), s))
    return "; ".join(ordered)


def build_summary_text(model: Model) -> str:
    lines = []
    lines.append(f"Μέγιστη ετήσια χωρητικότητα ανά έτος: {MAX_YEARLY_CAPACITY} ανθρωπομήνες\n")
    lines.append("Ετήσια σύνολα ανθρωπομηνών:\n")
    for y in sorted(model.yearly_totals.keys()):
        total = model.yearly_totals.get(y, 0)
        if total >= MAX_YEARLY_CAPACITY:
            lines.append(f"Έτος {y}: {total} (Η χωρητικότητα επιτεύχθηκε)")
        else:
            lines.append(f"Έτος {y}: {total}")
    lines.append("")

    deficits = []
    for r in sorted(model.row_period.keys()):
        req = model.row_requested_am.get(r, 0)
        alloc = model.row_x_count.get(r, 0)
        if req > alloc:
            deficits.append((r, req, alloc, req - alloc))

    if deficits:
        lines.append("Έργα με μη κατανεμημένους ανθρωπομήνες:\n")
        for r, req, alloc, unalloc in deficits:
            period = model.row_period.get(r, "")
            lines.append(f"Περίοδος: {period}, Αρχικοί ΑΜ: {req}, Κατανεμημένοι ΑΜ: {alloc}, Μη κατανεμημένοι ΑΜ: {unalloc}")
            reasons = compute_unallocated_reasons(model, r)
            if reasons:
                lines.append(f"Λόγοι μη κατανομής: {reasons}")
            lines.append("")
    else:
        lines.append("Δεν υπάρχουν έργα με μη κατανεμημένους ανθρωπομήνες.\n")

    return "\n".join(lines).strip()


def copy_style(style_obj):
    import copy
    return copy.copy(style_obj)


def process_excel(input_bytes: bytes) -> Tuple[bytes, str]:
    wb_in = openpyxl.load_workbook(io.BytesIO(input_bytes))
    ws_in = wb_in.active

    headers = {str(ws_in.cell(1, c).value).strip(): c for c in range(1, ws_in.max_column + 1)}
    if "ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ" not in headers or "ΑΝΘΡΩΠΟΜΗΝΕΣ" not in headers:
        raise RuntimeError("Το input πρέπει να έχει στήλες: ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ και ΑΝΘΡΩΠΟΜΗΝΕΣ")

    PERIOD_COL_IN = headers["ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ"]
    AM_COL_IN = headers["ΑΝΘΡΩΠΟΜΗΝΕΣ"]

    rows = []
    all_months = set()

    for r in range(2, ws_in.max_row + 1):
        period_cell = ws_in.cell(r, PERIOD_COL_IN)
        period_val = period_cell.value
        am_raw = ws_in.cell(r, AM_COL_IN).value

        try:
            am = int(am_raw) if am_raw is not None else 0
        except Exception:
            am = 0

        if not period_val or am == 0:
            continue

        start, end = parse_period(str(period_val))
        months = month_range(start, end)
        for ym in months:
            all_months.add(ym)

        rgb = period_cell.fill.start_color.rgb if period_cell.fill.start_color else None
        is_yellow = rgb in YELLOW_RGB_CANDIDATES

        rows.append({
            "period_str": str(period_val).strip(),
            "requested_am": am,
            "months_set": set(months),
            "months_list": sorted(months),
            "is_yellow": is_yellow,
        })

    if not rows:
        raise RuntimeError("Δεν βρέθηκαν έγκυρες γραμμές (με διάστημα και ΑΜ > 0).")

    years = sorted(set(y for y, _ in all_months))
    rows.sort(key=lambda x: (not x["is_yellow"], len(x["months_list"])))

    template_bytes = get_template_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb.active
    ws.freeze_panes = "D1"
    ws.title = "ΑΝΑΛΥΣΗ"

    cv_sheet = wb.create_sheet(title="CV", index=0)
    for row_idx, row_data in enumerate(ws_in.iter_rows()):
        for col_idx, cell in enumerate(row_data):
            new_cell = cv_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell.value)
            if cell.has_style:
                new_cell.font = copy_style(cell.font)
                new_cell.border = copy_style(cell.border)
                new_cell.fill = copy_style(cell.fill)
                new_cell.number_format = cell.number_format

    for col_idx in range(1, ws_in.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if col_letter in ws_in.column_dimensions:
            cv_sheet.column_dimensions[col_letter].width = ws_in.column_dimensions[col_letter].width

    cv_sheet["A1"] = "Α/Α"
    cv_sheet["A1"].font = Font(bold=True)
    cv_sheet["A1"].border = thin_border

    last_row_b = 0
    for rr in range(1, cv_sheet.max_row + 1):
        if cv_sheet.cell(rr, 2).value is not None:
            last_row_b = rr
    for i in range(2, last_row_b + 1):
        cv_sheet.cell(i, 1).value = i - 1
        cv_sheet.cell(i, 1).border = thin_border

    START_ROW_DATA = 4
    YEAR_ROW = 2
    MONTH_ROW = 3
    YEARLY_TOTAL_ROW = START_ROW_DATA + 1
    START_COL = 5

    for rng in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(str(rng))
        if (min_row <= YEAR_ROW <= max_row) or (min_row <= MONTH_ROW <= max_row) or (min_row <= START_ROW_DATA <= max_row) or (min_row <= YEARLY_TOTAL_ROW <= max_row):
            ws.unmerge_cells(str(rng))

    max_col_to_clear = max(START_COL + len(years) * 12, ws.max_column + 1)
    for r_clear in [YEAR_ROW, MONTH_ROW, START_ROW_DATA, YEARLY_TOTAL_ROW]:
        for c_clear in range(1, max_col_to_clear):
            ws.cell(r_clear, c_clear).value = None
            ws.cell(r_clear, c_clear).fill = PatternFill()

    for r_clear in range(START_ROW_DATA + 2, ws.max_row + 1):
        for c_clear in range(1, max_col_to_clear):
            ws.cell(r_clear, c_clear).value = None
            ws.cell(r_clear, c_clear).fill = PatternFill()

    month_col_map, end_col = build_headers_and_maps(ws, years, START_COL, YEAR_ROW, MONTH_ROW)
    ws.cell(YEARLY_TOTAL_ROW, 2).value = "ΕΤΗΣΙΑ ΣΥΝΟΛΑ"
    ws.cell(YEARLY_TOTAL_ROW, 2).font = Font(bold=True)
    ws.cell(YEARLY_TOTAL_ROW, 2).border = thin_border

    ws["A5"] = "Α/Α"
    ws["B2"] = "ΑΝΘΡΩΠΟΜΗΝΕΣ ΕΜΠΕΙΡΙΑΣ"
    ws["B2"].fill = orange_fill

    for c in range(START_COL, end_col):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 2.5

    model = Model()
    for ym, c in month_col_map.items():
        model.ym_to_col[ym] = c
        model.col_to_ym[c] = ym
    for y in years:
        model.year_blocks.append((y, month_col_map[(y, 1)], month_col_map[(y, 12)]))

    current_row = START_ROW_DATA + 2
    for idx, row in enumerate(rows, start=1):
        r = current_row
        model.row_period[r] = row["period_str"]
        model.row_months[r] = row["months_set"]
        model.row_months_list[r] = row["months_list"]
        model.row_requested_am[r] = int(row["requested_am"])
        model.row_is_yellow[r] = bool(row["is_yellow"])
        model.row_project_index[r] = idx
        model.row_x_count[r] = 0

        ws.cell(r, 1).value = f"=MATCH(B{r},CV!$B$2:$B${last_row_b},0)"
        ws.cell(r, 1).border = thin_border

        ws.cell(r, 2).value = model.row_period[r]
        ws.cell(r, 2).border = thin_border
        if model.row_is_yellow[r]:
            ws.cell(r, 2).fill = yellow_fill

        ws.cell(r, 3).value = model.row_requested_am[r]
        ws.cell(r, 3).border = thin_border

        for ym in model.row_months_list[r]:
            cell = ws.cell(r, model.ym_to_col[ym])
            cell.fill = yellow_fill
            cell.border = thin_border

        need = model.row_requested_am[r]
        got = 0
        for ym in model.row_months_list[r]:
            if got >= need:
                break
            if model.yearly_totals.get(ym[0], 0) >= MAX_YEARLY_CAPACITY:
                continue
            if ym in model.month_owner:
                continue
            ws.cell(r, model.ym_to_col[ym]).value = "X"
            model.month_owner[ym] = r
            model.yearly_totals[ym[0]] = model.yearly_totals.get(ym[0], 0) + 1
            model.row_x_count[r] += 1
            got += 1

        for c in range(START_COL, end_col):
            ws.cell(r, c).border = thin_border
        current_row += 1

    # Rule 3: ensure >=1 X
    for r in list(model.row_period.keys()):
        if model.row_requested_am.get(r, 0) > 0 and model.row_x_count.get(r, 0) == 0:
            for ym in model.row_months_list.get(r, []):
                if ym in model.month_owner:
                    continue
                if model.yearly_totals.get(ym[0], 0) >= MAX_YEARLY_CAPACITY:
                    if not free_capacity_in_year(model, ws, ym[0]):
                        continue
                if add_x(model, ws, r, ym):
                    break

    def optimize_rows(target_rows: List[int]) -> None:
        changed = True
        iters = 0
        while changed and iters < 6000:
            iters += 1
            changed = False
            deficits = [rr for rr in target_rows if model.row_requested_am.get(rr, 0) > 0 and model.row_x_count.get(rr, 0) < model.row_requested_am.get(rr, 0)]
            if not deficits:
                break
            deficits.sort(key=lambda rr: (-(model.row_requested_am[rr] - model.row_x_count[rr]), len(model.row_months_list.get(rr, []))))
            for rr in deficits:
                for ym in model.row_months_list.get(rr, []):
                    if ym in model.month_owner and model.month_owner[ym] != rr:
                        if not make_month_free(model, ws, ym, depth=0, max_depth=6, visited=set()):
                            continue
                    if ym in model.month_owner:
                        continue
                    if model.yearly_totals.get(ym[0], 0) >= MAX_YEARLY_CAPACITY:
                        if not free_capacity_in_year(model, ws, ym[0]):
                            continue
                    if add_x(model, ws, rr, ym):
                        changed = True
                        break
                if changed:
                    break

    yellow_rows = [r for r in model.row_period if model.is_yellow_row(r)]
    white_rows = [r for r in model.row_period if not model.is_yellow_row(r)]
    optimize_rows(yellow_rows)
    optimize_rows(white_rows)

    # totals row
    for y in years:
        start_c = month_col_map[(y, 1)]
        end_c = month_col_map[(y, 12)]
        if start_c != end_c:
            ws.merge_cells(start_row=YEARLY_TOTAL_ROW, start_column=start_c, end_row=YEARLY_TOTAL_ROW, end_column=end_c)
        total = model.yearly_totals.get(y, 0)
        cell = ws.cell(YEARLY_TOTAL_ROW, start_c)
        cell.value = total
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if total >= MAX_YEARLY_CAPACITY:
            ws.cell(YEAR_ROW, start_c).fill = red_fill
            ws.cell(YEAR_ROW, start_c).font = Font(color="FFFFFF", bold=True)
            cell.fill = red_fill
            cell.font = Font(color="FFFFFF", bold=True)
        elif total > 0:
            cell.fill = green_fill
            cell.font = Font(color="000000", bold=True)

    total_yellow_alloc = sum(model.row_x_count.get(r, 0) for r in yellow_rows)
    ws["C2"].value = total_yellow_alloc
    ws["C2"].font = Font(bold=True)
    ws["C2"].border = thin_border
    ws["C2"].fill = orange_fill

    ws["A6"] = f"=MATCH(B6,CV!$B$2:$B${last_row_b},0)"

    # red iff alloc < requested
    for r in model.row_period:
        req = model.row_requested_am.get(r, 0)
        alloc = model.row_x_count.get(r, 0)
        if req > 0 and alloc < req:
            ws.cell(r, 3).font = Font(color="FF0000", bold=True)
        else:
            ws.cell(r, 3).font = Font(color="000000", bold=False)

    summary = build_summary_text(model)
    out_buf = io.BytesIO()
    wb.save(out_buf)
    return out_buf.getvalue(), summary


# ------------------------- Sidebar: Usage Guide (sticky bottom-left) -------------------------
with st.sidebar:
    st.markdown('<div class="sidebar-bottom">', unsafe_allow_html=True)
    st.link_button("📘 Οδηγίες χρήσης της εφαρμογής", "?page=guide", use_container_width=True)
    st.download_button(
        label="⬇️ Λήψη οδηγιών (Word)",
        data=get_guide_doc_bytes(),
        file_name="Οδηγίες χρήσης Εργαλείου Κατανομής ΑΜ.docx",
        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        use_container_width=True,
    )
    st.markdown("</div>", unsafe_allow_html=True)


# ------------------------- Routing: main vs guide view -------------------------
page = st.query_params.get("page", "")
if page == "guide":
    left, right = st.columns([0.72, 0.28], vertical_alignment="center")
    with left:
        st.markdown("## Εργαλείο Κατανομής Ανθρωπομηνών")
        st.markdown("### Οδηγίες χρήσης της εφαρμογής")
    with right:
        try:
            st.image(get_logo_bytes(), width=LOGO_WIDTH_PX)
        except Exception:
            pass

    st.markdown("---")
    try:
        guide_text = get_guide_text()
        if guide_text:
            st.markdown(guide_text.replace("\n", "  \n"))
        else:
            st.info("Δεν βρέθηκε κείμενο στο αρχείο οδηγιών.")
    except Exception as e:
        st.error(f"Αποτυχία φόρτωσης οδηγιών: {e}")

    st.markdown("---")
    st.link_button("⬅️ Επιστροφή στην εφαρμογή", "?", use_container_width=False)
    st.stop()


# ------------------------- Main page -------------------------
left, right = st.columns([0.72, 0.28], vertical_alignment="center")
with left:
    st.markdown("## Εργαλείο Κατανομής Ανθρωπομηνών")
    st.markdown("Αυτό το εργαλείο κατανέμει ανθρωπομήνες σε έργα με βάση χρονικά διαστήματα και μέγιστη ετήσια χωρητικότητα.")
with right:
    try:
        st.image(get_logo_bytes(), width=LOGO_WIDTH_PX)
    except Exception:
        pass

st.markdown("---")
st.markdown("### 👉 Ανέβασε το INPUT excel (μόνο 2 στήλες)")

uploaded = st.file_uploader(
    " ",
    type=["xlsx"],
    accept_multiple_files=False,
    help="Το Excel πρέπει να έχει στήλες: ΧΡΟΝΙΚΟ ΔΙΑΣΤΗΜΑ και ΑΝΘΡΩΠΟΜΗΝΕΣ.",
)

if not uploaded:
    st.markdown('<div class="notice">Παρακαλώ ανεβάστε ένα αρχείο Excel για να ξεκινήσετε την επεξεργασία.</div>', unsafe_allow_html=True)
    st.stop()

st.write(f"**Επιλεγμένο αρχείο:** `{uploaded.name}`")
run_btn = st.button("✅ Εκτέλεση κατανομής", use_container_width=True)

if run_btn:
    with st.spinner("Εκτελείται η κατανομή..."):
        try:
            out_bytes, summary_text = process_excel(uploaded.getvalue())
            st.session_state["out_bytes"] = out_bytes
            st.session_state["summary_text"] = summary_text
            st.session_state["out_name"] = os.path.splitext(uploaded.name)[0] + "_ΚΑΤΑΝΟΜΗ ΑΜ.xlsx"
            st.markdown('<div class="successbox">Το αρχείο επεξεργάστηκε επιτυχώς!</div>', unsafe_allow_html=True)
        except Exception as e:
            st.session_state.pop("out_bytes", None)
            st.session_state.pop("summary_text", None)
            st.markdown(f'<div class="errorbox">Αποτυχία επεξεργασίας: {e}</div>', unsafe_allow_html=True)

if "out_bytes" in st.session_state:
    st.download_button(
        label="⬇️ Κατεβάστε το επεξεργασμένο Excel",
        data=st.session_state["out_bytes"],
        file_name=st.session_state.get("out_name", "output.xlsx"),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.markdown("### **Σύνοψη Κατανομής**")
    st.text(st.session_state.get("summary_text", ""))
